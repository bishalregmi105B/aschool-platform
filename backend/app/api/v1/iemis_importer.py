"""IEMIS Data Importer API.

Nepal Ministry of Education IEMIS (Integrated Education Management Information System)
exports data as Excel files. This API parses those exports and maps them to
ASchool's data models.

Supported formats:
  - student_namewise  : Student Name-wise Report (per-student row with parent info)
  - school_level      : School Level Report (one-row-per-school metadata)

Endpoints:
  POST /iemis/validate         — Validate Excel file and preview rows (dry run)
  POST /iemis/import           — Live import into the database
  GET  /iemis/history          — List past imports for this school
  GET  /iemis/history/<id>     — Detail of a past import (rows, errors)
  GET  /iemis/formats          — List supported IEMIS formats with column maps
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from typing import Any

from flask import Blueprint, g, request
from flask_jwt_extended import get_jwt_identity, jwt_required

from app.plugins.decorators import plugin_required
from app.plugins.entitlements import StudentCapExceededError
from app.services.student_numbers import ensure_student_numbers
from app.utils.decorators import role_required, school_required
from app.utils.response import created_response, error_response, success_response
from extensions import db

iemis_importer_bp = Blueprint("iemis_importer", __name__, url_prefix="/iemis")

# ── Import Job model ──────────────────────────────────────────────────────────
from app.models.iemis import IemisImportLog  # noqa: E402

# ── Column Mappings ───────────────────────────────────────────────────────────

STUDENT_NAMEWISE_COLUMNS = {
    "S.N": "sn",
    "IEMIS Code": "iemis_school_code",
    "Current School": "school_name_iemis",
    "Student Id": "iemis_student_id",
    "Full Name": "full_name",
    "Gender": "gender",
    "Class": "grade",
    "DOB": "dob",
    "Age": "age",
    "Father Name": "father_name",
    "Mother Name": "mother_name",
    "Guardian Name": "guardian_name",
    "Guardian Contact Number": "guardian_phone",
    "Section": "section",
    "Permanent Address": "permanent_address",
    "Temporary Address": "temporary_address",
}

SCHOOL_LEVEL_COLUMNS = {
    "S.N": "sn",
    "Iemis Code": "iemis_code",
    "School Name": "name",
    "School Type": "school_type",
    "School Sub Type": "school_sub_type",
    "Province": "province",
    "District": "district",
    "Municipality": "municipality",
    "Ward": "ward",
    "Tole": "tole",
    "Head Teacher Name": "head_teacher_name",
    "Head Teacher Contact Number": "head_teacher_phone",
    "School Email": "email",
    "Eced Establishment Date": "eced_established",
    "Basic Level (1–5) Establishment Date": "basic_1_5_established",
    "Basic Level (6–8) Establishment Date": "basic_6_8_established",
    "Secondary Level (9–10) Establishment Date": "secondary_9_10_established",
    "Secondary Level (11–12) Establishment Date": "secondary_11_12_established",
    "SEE Code": "see_code",
    "HSEB Code": "hseb_code",
    "Class Registered Upto": "max_class",
    "School Level": "school_level",
}

STAFF_DETAILS_COLUMNS = {
    "S.N": "sn",
    "Iemis Code": "iemis_code",
    "Teacher Id": "iemis_teacher_id",
    "Full Name": "full_name",
    "Gender": "gender",
    "DOB": "dob",
    "Contact Number": "phone",
    "Email": "email",
    "Designation": "designation",
    "Level": "level",
    "Appointment Status": "appointment_status",
    "Teaching Subject": "teaching_subject",
}

FORMAT_MAP = {
    "student_namewise": {
        "name": "Student Name-wise Report",
        "sheet": "Student_Namewise_Report",
        "columns": STUDENT_NAMEWISE_COLUMNS,
    },
    "school_level": {
        "name": "School Level Report",
        "sheet": "School_Level_Report",
        "columns": SCHOOL_LEVEL_COLUMNS,
    },
    "staff_details": {
        "name": "Staff Details Report",
        "sheet": "Staff_Details",
        "columns": STAFF_DETAILS_COLUMNS,
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────────


def _parse_excel(file_bytes: bytes, format_code: str) -> tuple[list[dict], list[str]]:
    """Parse IEMIS Excel export. Returns (rows_as_dicts, warnings)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    fmt = FORMAT_MAP.get(format_code)
    if not fmt:
        raise ValueError(f"Unknown format: {format_code}")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Try configured sheet name first, then first sheet
    sheet_name = fmt["sheet"]
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    col_map = fmt["columns"]
    warnings = []
    rows_iter = ws.iter_rows(values_only=True)

    # Find header row (first non-empty row)
    headers = None
    for row in rows_iter:
        non_empty = [c for c in row if c is not None]
        if non_empty:
            headers = [str(c).strip() if c else "" for c in row]
            break

    if not headers:
        raise ValueError("Excel file appears empty — no header row found")

    # Map header → field name
    field_map: dict[int, str] = {}
    for idx, h in enumerate(headers):
        if h in col_map:
            field_map[idx] = col_map[h]
        else:
            warnings.append(f"Unknown column '{h}' — ignored")

    parsed_rows = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        record: dict[str, Any] = {}
        for idx, val in enumerate(row):
            fname = field_map.get(idx)
            if fname:
                record[fname] = val
        parsed_rows.append(record)

    wb.close()
    return parsed_rows, warnings


def _parse_csv(file_bytes: bytes, format_code: str) -> tuple[list[dict], list[str]]:
    """Parse a CSV export using the same FORMAT_MAP headers as the Excel path.

    Added for the Generic CSV Upload page (bulk-uploads/csv): the frontend
    previously faked the upload because only .xlsx was accepted here.
    """
    import csv as _csv

    fmt = FORMAT_MAP.get(format_code)
    if not fmt:
        raise ValueError(f"Unknown format: {format_code}")

    col_map = fmt["columns"]
    warnings: list[str] = []

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = _csv.reader(io.StringIO(text))
    headers = None
    for row in reader:
        if any(str(c).strip() for c in row):
            headers = [str(c).strip() if c else "" for c in row]
            break

    if not headers:
        raise ValueError("CSV file appears empty — no header row found")

    field_map: dict[int, str] = {}
    for idx, h in enumerate(headers):
        if h in col_map:
            field_map[idx] = col_map[h]
        else:
            warnings.append(f"Unknown column '{h}' — ignored")

    parsed_rows: list[dict[str, Any]] = []
    for row in reader:
        if not row or all(not str(c).strip() for c in row):
            continue
        record: dict[str, Any] = {}
        for idx, val in enumerate(row):
            fname = field_map.get(idx)
            if fname and str(val).strip():
                record[fname] = val.strip()
        parsed_rows.append(record)

    return parsed_rows, warnings


def _parse_tabular(file_bytes: bytes, ext: str, format_code: str) -> tuple[list[dict], list[str]]:
    """Dispatch parsing by file extension — csv or excel."""
    if ext == "csv":
        return _parse_csv(file_bytes, format_code)
    return _parse_excel(file_bytes, format_code)


def _gender_normalize(val: Any) -> str:
    if not val:
        return "other"
    v = str(val).strip().upper()
    if v in ("M", "MALE", "BOY", "पुरुष"):
        return "male"
    if v in ("F", "FEMALE", "GIRL", "महिला"):
        return "female"
    return "other"


def _safe_str(val: Any, max_len: int = 300) -> str | None:
    if val is None:
        return None
    return str(val).strip()[:max_len] or None


# users.phone is NOT NULL (String(20), models/user.py:44) and there is no
# unique constraint on it, but parent linking looks users up by
# (school_id, phone) — so placeholders must be deterministic per source row
# and collision-free within the school.
PLACEHOLDER_PHONE_PREFIX = "9800000"  # reserved 9800000000–9800999999 block


def _placeholder_phone(school_id, unique_key: str) -> str:
    """Deterministic Nepal-format placeholder phone for an imported user whose
    source row carries no phone number (users.phone is NOT NULL).

    Format: 9800000xxxx (11 chars). The 4-digit suffix is derived from a keyed
    hash of (school_id, unique_key) so re-imports of the same row yield the
    same number, then bumped deterministically until free within the school to
    avoid stealing a real parent's lookup phone. Callers must mark the user
    (permissions["placeholder_phone"]=True, phone_verified stays False) so the
    number is never mistaken for a real contact.
    """
    import hashlib

    from app.models.user import User

    digest = hashlib.blake2b(
        f"{school_id}:{unique_key}".encode("utf-8"), digest_size=8
    ).digest()
    base = int.from_bytes(digest, "big") % 10000
    for offset in range(10000):
        candidate = f"{PLACEHOLDER_PHONE_PREFIX}{(base + offset) % 10000:04d}"
        taken = User.query.filter_by(school_id=school_id, phone=candidate).first()
        if not taken:
            return candidate
    # 10k suffixes exhausted for this school (implausible) — fall back to a
    # collision-safe row-unique number; uniqueness beats realism here.
    return f"{PLACEHOLDER_PHONE_PREFIX}{int.from_bytes(digest, 'big') % 100000000:08d}"[:20]


def _parse_dob(val: Any) -> tuple[str | None, date | None]:
    """Return (dob_bs_string, dob_ad_date). IEMIS uses BS dates as strings like 2071-07-28."""
    if not val:
        return None, None
    s = str(val).strip()
    # IEMIS stores BS date as YYYY-MM-DD string
    dob_bs = s if len(s) == 10 and s[4] == "-" else None
    return dob_bs, None  # AD conversion would need nepalicalendar lib


def _normalize_grade(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    low = s.lower()
    if low.startswith("class "):
        s = s[6:].strip()
    return s[:10] or None


# ── Import Logic ──────────────────────────────────────────────────────────────


def _import_students(rows: list[dict], school_id, dry_run: bool = False) -> dict:
    """Map IEMIS student rows to ASchool Student + Guardian records."""
    from app.models.school import School
    from app.models.student import Guardian, Student
    from app.models.user import User
    from app.utils.password import generate_default_password

    try:
        from app.models.academic import Class, Section

        has_class_model = True
    except ImportError:
        has_class_model = False

    school_obj = School.query.get(school_id)

    def _resolve_or_create_parent_user(
        name: str | None, phone: str | None
    ) -> User | None:
        phone_clean = (phone or "").strip()
        if not phone_clean:
            return None

        existing = User.query.filter_by(
            school_id=school_id,
            phone=phone_clean,
            is_deleted=False,
        ).first()
        if existing:
            if existing.role != "parent":
                return None
            if name and existing.full_name != name:
                existing.full_name = name
            return existing

        parent_user = User(
            school_id=school_id,
            role="parent",
            full_name=name or "Parent/Guardian",
            phone=phone_clean,
        )
        parent_user.set_password(
            generate_default_password(parent_user, None, school_obj)
        )
        db.session.add(parent_user)
        db.session.flush()
        return parent_user

    imported = skipped = errors_count = 0
    error_list = []
    previews = []

    # ── Server-side plan cap (E2): reject the import when the new students it
    #    would create would push the school past School.max_students
    #    (NULL/0 = unlimited). Preview/dry-run imports are not capped.
    if not dry_run:
        from app.plugins.entitlements import (
            StudentCapExceededError,
            student_cap_error,
        )

        iemis_ids = [
            iid for iid in (_safe_str(r.get("iemis_student_id"), 100) for r in rows) if iid
        ]
        existing_iemis_ids: set = set()
        if iemis_ids:
            existing_iemis_ids = {
                s.student_id
                for s in Student.query.filter(
                    Student.school_id == school_id,
                    Student.student_id.in_(iemis_ids),
                    Student.is_deleted.is_(False),
                ).all()
            }
        prospective_new = sum(
            1
            for row in rows
            if _safe_str(row.get("full_name"))  # nameless rows never create students
            and (
                not _safe_str(row.get("iemis_student_id"), 100)
                or _safe_str(row.get("iemis_student_id"), 100) not in existing_iemis_ids
            )
        )
        cap_error = student_cap_error(school_id, incoming_count=prospective_new)
        if cap_error:
            raise StudentCapExceededError(cap_error)

    for i, row in enumerate(rows, start=2):  # row 2 = first data row
        full_name = _safe_str(row.get("full_name"))
        if not full_name:
            errors_count += 1
            error_list.append({"row": i, "error": "Missing required field: Full Name"})
            continue

        # Split name naively: last word is last name
        parts = full_name.split()
        first_name = " ".join(parts[:-1]) if len(parts) > 1 else full_name
        last_name = parts[-1] if len(parts) > 1 else ""

        iemis_id = _safe_str(row.get("iemis_student_id"), 100)
        gender = _gender_normalize(row.get("gender"))
        grade = _normalize_grade(row.get("grade"))
        section = _safe_str(row.get("section"), 50)
        dob_bs, dob_ad = _parse_dob(row.get("dob"))
        perm_address = _safe_str(row.get("permanent_address"))
        temp_address = _safe_str(row.get("temporary_address"))

        guardian_name = _safe_str(row.get("guardian_name"), 200)
        guardian_phone = _safe_str(row.get("guardian_phone"), 20)
        father_name = _safe_str(row.get("father_name"), 200)
        mother_name = _safe_str(row.get("mother_name"), 200)

        preview = {
            "row": i,
            "full_name": full_name,
            "first_name": first_name,
            "last_name": last_name,
            "iemis_student_id": iemis_id,
            "gender": gender,
            "grade": grade,
            "section": section,
            "dob_bs": dob_bs,
            "guardian_name": guardian_name,
            "guardian_phone": guardian_phone,
            "father_name": father_name,
            "mother_name": mother_name,
            "permanent_address": perm_address,
        }
        previews.append(preview)

        if dry_run:
            imported += 1
            continue

        try:
            with db.session.begin_nested():
                # Check if student already exists by IEMIS ID
                existing = None
                if iemis_id:
                    existing = Student.query.filter_by(
                        school_id=school_id,
                        student_id=iemis_id,
                        is_deleted=False,
                    ).first()

                # Auto-link or create Class/Section
                class_id = None
                section_id = None
                if has_class_model and grade:
                    klass = Class.query.filter_by(
                        school_id=school_id, name=grade, is_deleted=False
                    ).first()
                    if not klass:
                        klass = Class(school_id=school_id, name=grade)
                        db.session.add(klass)
                        db.session.flush()
                    class_id = klass.id

                    if section:
                        sec = Section.query.filter_by(
                            school_id=school_id,
                            class_id=class_id,
                            name=section,
                            is_deleted=False,
                        ).first()
                        if not sec:
                            sec = Section(
                                school_id=school_id, class_id=class_id, name=section
                            )
                            db.session.add(sec)
                            db.session.flush()
                        section_id = sec.id

                if existing:
                    # Update existing record
                    existing.first_name = first_name
                    existing.last_name = last_name
                    if iemis_id:
                        existing.student_id = iemis_id
                    if gender:
                        existing.gender = gender
                    if dob_bs:
                        existing.dob_bs = dob_bs
                    if class_id:
                        existing.class_id = class_id
                    if section_id:
                        existing.section_id = section_id
                    if grade:
                        existing.academic_year = grade

                    existing_addr = dict(existing.address or {})
                    if perm_address:
                        existing_addr["permanent"] = perm_address
                    if temp_address:
                        existing_addr["temporary"] = temp_address
                    if existing_addr:
                        existing.address = existing_addr

                    student = existing
                    # E235: re-imports of pre-existing rows also backfill a
                    # missing enrollment number / roll ("if not present").
                    ensure_student_numbers(student)
                    skipped += 1
                else:
                    student = Student(
                        school_id=school_id,
                        first_name=first_name,
                        last_name=last_name,
                        student_id=iemis_id,
                        gender=gender,
                        dob_bs=dob_bs,
                        dob_ad=dob_ad,
                        academic_year=grade,
                        class_id=class_id,
                        section_id=section_id,
                        address={
                            "permanent": perm_address,
                            "temporary": temp_address,
                        },
                    )
                    # E235: IEMIS rows carry the external EMIS id in
                    # student_id; the school's own enrollment number and
                    # class roll are auto-assigned here when missing
                    # (School-row FOR UPDATE lock serializes issuance across
                    # concurrent imports/creations).
                    ensure_student_numbers(student)
                    db.session.add(student)
                    db.session.flush()  # get student.id

                    # Generate a unique internal email for student login.
                    # Format: stud.<iemis_id>@<school_slug>.import.local
                    # This allows login via email without real email data.
                    school_slug = getattr(school_obj, "slug", "school") or "school"
                    if iemis_id:
                        student_email = f"stud.{iemis_id}@{school_slug}.import.local"
                    else:
                        student_email = (
                            f"stud.{school_id}.{i}@{school_slug}.import.local"
                        )

                    # Create User account for student.
                    # users.phone is NOT NULL — generate a deterministic
                    # placeholder (9800000xxxx) when the IEMIS row has no
                    # contact number, and mark the user so it is never
                    # mistaken for a real one.
                    student_phone = _placeholder_phone(
                        school_id, f"stud:{iemis_id or i}:{full_name}"
                    )
                    user = User(
                        school_id=school_id,
                        role="student",
                        full_name=f"{first_name} {last_name}".strip(),
                        phone=student_phone,
                        email=student_email,
                        permissions={"placeholder_phone": True},
                    )
                    user.set_password(
                        generate_default_password(user, student, school_obj)
                    )

                    db.session.add(user)
                    db.session.flush()
                    student.user_id = user.id
                    imported += 1

                # Upsert guardian records to avoid duplicates across repeated imports.
                guardian_records = Guardian.query.filter_by(
                    school_id=school_id,
                    student_id=student.id,
                    is_deleted=False,
                ).all()

                def _upsert_guardian(
                    relation: str,
                    name: str | None,
                    phone: str | None = None,
                    is_primary: bool = False,
                ):
                    if not name and not phone:
                        return
                    guardian = next(
                        (g for g in guardian_records if g.relation == relation), None
                    )
                    if not guardian and name:
                        guardian = next(
                            (g for g in guardian_records if g.full_name == name), None
                        )
                    if not guardian:
                        guardian = Guardian(
                            school_id=school_id,
                            student_id=student.id,
                            relation=relation,
                            full_name=name or "Unknown",
                        )
                        db.session.add(guardian)
                        guardian_records.append(guardian)
                    if name:
                        guardian.full_name = name
                    if phone:
                        guardian.phone = phone
                    guardian.relation = relation
                    if is_primary:
                        guardian.is_primary = True

                    parent_user = _resolve_or_create_parent_user(name, phone)
                    if parent_user:
                        guardian.user_id = parent_user.id

                _upsert_guardian(
                    "guardian",
                    guardian_name or father_name or mother_name,
                    phone=guardian_phone,
                    is_primary=True,
                )
                if father_name and father_name != guardian_name:
                    _upsert_guardian("father", father_name)
                if mother_name and mother_name != guardian_name:
                    _upsert_guardian("mother", mother_name)

        except Exception as exc:
            errors_count += 1
            error_name = full_name if "full_name" in locals() else "Unknown Row"
            error_list.append(
                {"row": i, "name": error_name, "error": f"Import failed: {str(exc)}"}
            )
            continue

    if not dry_run:
        db.session.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors_count,
        "error_list": error_list[:50],  # cap at 50 for response
        "previews": previews[:20] if dry_run else [],
    }


def _import_school_level(rows: list[dict], school_id, dry_run: bool = False) -> dict:
    """Map IEMIS school-level rows to ASchool School model."""
    from app.models.school import School

    imported = skipped = errors_count = 0
    error_list = []
    previews = []

    for i, row in enumerate(rows, start=2):
        iemis_code = _safe_str(row.get("iemis_code"), 50)
        school_name = _safe_str(row.get("name"), 300)

        if not iemis_code and not school_name:
            errors_count += 1
            error_list.append(
                {"row": i, "error": "Missing both IEMIS Code and School Name"}
            )
            continue

        preview = {
            "row": i,
            "iemis_code": iemis_code,
            "name": school_name,
            "province": _safe_str(row.get("province")),
            "district": _safe_str(row.get("district")),
            "municipality": _safe_str(row.get("municipality")),
            "head_teacher": _safe_str(row.get("head_teacher_name")),
            "email": _safe_str(row.get("email")),
            "school_level": _safe_str(row.get("school_level")),
        }
        previews.append(preview)

        if dry_run:
            imported += 1
            continue

        try:
            school = School.query.filter_by(id=school_id).first()
            if school:
                from sqlalchemy import inspect as sa_inspect

                col_names = {c.key for c in sa_inspect(School).mapper.column_attrs}

                # ── Update real School columns where they exist ──────────────
                if _safe_str(row.get("province")) and "province" in col_names:
                    school.province = _safe_str(row.get("province"), 100)
                if _safe_str(row.get("district")) and "district" in col_names:
                    school.district = _safe_str(row.get("district"), 100)
                if _safe_str(row.get("municipality")) and "municipality" in col_names:
                    school.municipality = _safe_str(row.get("municipality"), 100)
                if row.get("ward") and "ward" in col_names:
                    school.ward = str(row.get("ward"))[:10]
                if _safe_str(row.get("email")) and "email" in col_names:
                    if not school.email:  # only fill if blank
                        school.email = _safe_str(row.get("email"), 200)
                if _safe_str(row.get("head_teacher_phone")) and "phone" in col_names:
                    if not school.phone:
                        school.phone = _safe_str(row.get("head_teacher_phone"), 20)

                # ── Store IEMIS-specific metadata in settings JSONB ──────────
                meta = dict(school.settings or {})
                meta.update(
                    {
                        "iemis_code": iemis_code,
                        "iemis_school_type": _safe_str(row.get("school_type")),
                        "iemis_school_sub_type": _safe_str(row.get("school_sub_type")),
                        "iemis_see_code": _safe_str(row.get("see_code")),
                        "iemis_hseb_code": _safe_str(row.get("hseb_code")),
                        "iemis_max_class": _safe_str(row.get("max_class")),
                        "iemis_school_level": _safe_str(row.get("school_level")),
                        "iemis_head_teacher_name": _safe_str(
                            row.get("head_teacher_name")
                        ),
                        "iemis_head_teacher_phone": _safe_str(
                            row.get("head_teacher_phone")
                        ),
                    }
                )
                school.settings = meta
                from sqlalchemy.orm.attributes import flag_modified

                flag_modified(school, "settings")
                db.session.commit()
                imported += 1
            else:
                skipped += 1

        except Exception as exc:
            errors_count += 1
            error_list.append(
                {"row": i, "error": f"Failed to update school: {str(exc)}"}
            )
            db.session.rollback()

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors_count,
        "error_list": error_list[:50],
        "previews": previews[:20] if dry_run else [],
    }


def _import_staff(rows: list[dict], school_id, dry_run: bool = False) -> dict:
    """Map IEMIS staff details to ASchool User (staff) records."""
    from app.models.user import User
    from app.utils.password import generate_default_password

    imported = skipped = errors_count = 0
    error_list = []
    previews = []

    for i, row in enumerate(rows, start=2):
        try:
            full_name = _safe_str(row.get("full_name"))
            if not full_name:
                errors_count += 1
                error_list.append(
                    {"row": i, "error": "Missing required field: Full Name"}
                )
                continue

            iemis_id = _safe_str(row.get("iemis_teacher_id"), 100)
            designation = _safe_str(row.get("designation"), 100)
            phone = _safe_str(row.get("phone"), 20)
            email = _safe_str(row.get("email"), 150)
            gender = _gender_normalize(row.get("gender"))
            subject = _safe_str(row.get("teaching_subject"))

            preview = {
                "row": i,
                "full_name": full_name,
                "iemis_teacher_id": iemis_id,
                "designation": designation,
                "phone": phone,
                "email": email,
                "gender": gender,
                "teaching_subject": subject,
            }
            previews.append(preview)

            if dry_run:
                imported += 1
                continue

            # Check if teacher exists by iemis_id or email
            existing = None
            if email:
                existing = User.query.filter_by(
                    school_id=school_id, email=email, is_deleted=False
                ).first()
            if not existing and iemis_id:
                # Store IEMIS ID in settings or similar if we had a column, but for now we search by name/phone as fallback
                existing = User.query.filter_by(
                    school_id=school_id,
                    full_name=full_name,
                    phone=phone,
                    is_deleted=False,
                ).first()

            if existing:
                existing.full_name = full_name
                if phone:
                    existing.phone = phone
                if email:
                    existing.email = email
                if gender:
                    existing.gender = gender

                # User has no `settings` column — `permissions` (JSONB) is the
                # de-facto per-user settings bag (auth.py stores totp_secret /
                # mfa_enabled there), so IEMIS metadata lives under it too.
                meta = dict(existing.permissions or {})
                meta.update(
                    {
                        "iemis_teacher_id": iemis_id,
                        "iemis_designation": designation,
                        "iemis_teaching_subject": subject,
                    }
                )
                existing.permissions = meta
                skipped += 1
            else:
                # users.phone is NOT NULL — same deterministic placeholder rule
                # as the student import when the row has no contact number.
                staff_phone = phone or _placeholder_phone(
                    school_id, f"staff:{iemis_id or i}:{full_name}"
                )
                user_permissions = {} if phone else {"placeholder_phone": True}
                user_permissions.update(
                    {
                        "iemis_teacher_id": iemis_id,
                        "iemis_designation": designation,
                        "iemis_teaching_subject": subject,
                    }
                )
                user = User(
                    school_id=school_id,
                    role="teacher"
                    if "teacher" in str(designation).lower() or subject
                    else "staff",
                    full_name=full_name,
                    phone=staff_phone,
                    email=email,
                    gender=gender,
                    permissions=user_permissions,
                )
                from app.models.school import School

                school_obj = School.query.get(school_id)
                user.set_password(generate_default_password(user, None, school_obj))

                db.session.add(user)
                db.session.flush()  # get user.id before creating Staff record

                # Also create a Staff model record if the model supports it
                try:
                    from sqlalchemy import inspect as sa_inspect

                    from app.models.staff import Staff as StaffModel

                    # `Staff` is an alias of User right now — creating one would
                    # insert a bare second User row (no full_name/role) and fail
                    # the batch commit. Only proceed for a genuinely separate
                    # model with its own required columns.
                    if StaffModel is not User:
                        staff_cols = {
                            c.key for c in sa_inspect(StaffModel).mapper.column_attrs
                        }
                        staff_kwargs: dict = {
                            "school_id": school_id,
                            "user_id": user.id,
                        }
                        if "designation" in staff_cols and designation:
                            staff_kwargs["designation"] = designation
                        if "department" in staff_cols and subject:
                            staff_kwargs["department"] = subject
                        if "gender" in staff_cols and gender:
                            staff_kwargs["gender"] = gender
                        if "phone" in staff_cols and phone:
                            staff_kwargs["phone"] = phone
                        if "email" in staff_cols and email:
                            staff_kwargs["email"] = email
                        staff_record = StaffModel(**staff_kwargs)
                        db.session.add(staff_record)
                except (ImportError, Exception):
                    pass  # Staff model may have different schema; User record is sufficient

                imported += 1

        except Exception as exc:
            errors_count += 1
            error_name = full_name if "full_name" in locals() else "Unknown Row"
            error_list.append(
                {"row": i, "name": error_name, "error": f"Import failed: {str(exc)}"}
            )
            db.session.rollback()

    if not dry_run:
        db.session.commit()

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors_count,
        "error_list": error_list[:50],
        "previews": previews[:20] if dry_run else [],
    }


# ── Auto-detect format from headers ──────────────────────────────────────────


def _detect_format(file_bytes: bytes) -> str | None:
    """Try to auto-detect IEMIS format by reading the first row of the Excel."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        ws = wb.worksheets[0]
        for row in ws.iter_rows(max_row=3, values_only=True):
            non_empty = [str(c).strip() for c in row if c]
            if "Student Id" in non_empty or "Full Name" in non_empty:
                wb.close()
                return "student_namewise"
            if "Iemis Code" in non_empty and "School Name" in non_empty:
                wb.close()
                return "school_level"
            if "Teacher Id" in non_empty and "Designation" in non_empty:
                wb.close()
                return "staff_details"
        wb.close()
    except Exception:
        pass
    return None


# ── Routes ────────────────────────────────────────────────────────────────────


@iemis_importer_bp.route("/formats", methods=["GET"])
@jwt_required()
@school_required
@plugin_required("iemis_importer")
def list_formats():
    """Return supported IEMIS formats with column maps."""
    return success_response(
        [
            {
                "code": code,
                "name": fmt["name"],
                "sheet": fmt["sheet"],
                "columns": [
                    {"iemis_column": k, "aschool_field": v}
                    for k, v in fmt["columns"].items()
                ],
            }
            for code, fmt in FORMAT_MAP.items()
        ]
    )


@iemis_importer_bp.route("/template", methods=["GET"])
@jwt_required()
@school_required
@plugin_required("iemis_importer")
def download_template():
    """Download a ready-to-fill IEMIS import template (.xlsx).

    Generates a workbook whose header row exactly matches the IEMIS columns
    for the requested format, pre-filled with one honest sample row so the
    validate → import round-trip can be exercised end-to-end.
    """
    from flask import Response, send_file

    format_code = (request.args.get("format") or "student_namewise").strip()
    fmt = FORMAT_MAP.get(format_code)
    if not fmt:
        return error_response(
            f"Unknown format '{format_code}'. Supported: {', '.join(FORMAT_MAP)}", 400
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return error_response("openpyxl is required: pip install openpyxl", 500)

    wb = Workbook()
    ws = wb.active
    ws.title = fmt["sheet"][:31]  # Excel sheet-name limit

    columns = list(fmt["columns"].keys())
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # One honest sample row (values map to the first few ASchool fields).
    sample_values = {
        "gender": "Male",
        "grade": "10",
        "dob": "2065-04-15",
        "phone": "98XXXXXXXX",
        "email": "example@example.com",
        "school_level": "Secondary",
        "full_name": "Sample Student",
        "head_teacher_name": "Sample Principal",
    }
    ws.append([
        sample_values.get(fmt["columns"][col], f"Sample {fmt['columns'][col]}")
        for col in columns
    ])

    # Column widths sized to the header text
    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(
            max(len(col) + 2, 12), 40
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"iemis_template_{format_code}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@iemis_importer_bp.route("/validate", methods=["POST"])
@jwt_required()
@school_required
@plugin_required("iemis_importer")
@role_required("school_admin", "data_entry")
def validate_import():
    """Validate IEMIS Excel file and return a preview (dry run — no DB writes)."""
    if "file" not in request.files:
        return error_response(
            "No file uploaded. Send 'file' as multipart/form-data.", 400
        )

    f = request.files["file"]
    if not f.filename:
        return error_response("Empty filename", 400)

    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        return error_response("Only .xlsx, .xls and .csv files are supported", 415)

    file_bytes = f.read()
    if len(file_bytes) > 20 * 1024 * 1024:
        return error_response("File exceeds 20 MB limit", 413)

    format_code = request.form.get("format") or _detect_format(file_bytes)
    if not format_code:
        return error_response(
            "Cannot auto-detect format. Specify 'format' in form: student_namewise or school_level",
            400,
        )

    try:
        rows, warnings = _parse_tabular(file_bytes, ext, format_code)
    except (ValueError, RuntimeError) as exc:
        return error_response(str(exc), 422)

    if format_code == "student_namewise":
        result = _import_students(rows, g.school_id, dry_run=True)
    elif format_code == "staff_details":
        result = _import_staff(rows, g.school_id, dry_run=True)
    else:
        result = _import_school_level(rows, g.school_id, dry_run=True)

    return success_response(
        {
            "format": format_code,
            "filename": f.filename,
            "total_rows": len(rows),
            "valid_rows": result["imported"],
            "warnings": warnings,
            "preview": result.get("previews", []),
        }
    )


@iemis_importer_bp.route("/import", methods=["POST"])
@jwt_required()
@school_required
@plugin_required("iemis_importer")
@role_required("school_admin", "data_entry")
def run_import():
    """Execute the IEMIS import — parses file and writes to database."""
    if "file" not in request.files:
        return error_response(
            "No file uploaded. Send 'file' as multipart/form-data.", 400
        )

    f = request.files["file"]
    ext = (
        f.filename.rsplit(".", 1)[-1].lower()
        if f.filename and "." in f.filename
        else ""
    )
    if ext not in ("xlsx", "xls", "csv"):
        return error_response("Only .xlsx, .xls and .csv files are supported", 415)

    file_bytes = f.read()
    if len(file_bytes) > 20 * 1024 * 1024:
        return error_response("File exceeds 20 MB limit", 413)

    format_code = request.form.get("format") or _detect_format(file_bytes)
    if not format_code:
        return error_response(
            "Cannot auto-detect format. Specify 'format': student_namewise or school_level",
            400,
        )

    # Create import log
    log = IemisImportLog(
        school_id=g.school_id,
        imported_by=get_jwt_identity(),
        format_code=format_code,
        filename=f.filename,
        status="processing",
    )
    db.session.add(log)
    db.session.commit()

    try:
        rows, warnings = _parse_tabular(file_bytes, ext, format_code)
        log.total_rows = len(rows)

        if format_code == "student_namewise":
            result = _import_students(rows, g.school_id, dry_run=False)
        elif format_code == "staff_details":
            result = _import_staff(rows, g.school_id, dry_run=False)
        else:
            result = _import_school_level(rows, g.school_id, dry_run=False)

        log.imported_rows = result["imported"]
        log.skipped_rows = result["skipped"]
        log.error_rows = result["errors"]
        log.errors = result["error_list"]
        log.summary = {
            "warnings": warnings,
            "format": format_code,
        }
        log.status = "completed" if result["errors"] == 0 else "partial"
        log.completed_at = datetime.now(UTC)
        db.session.commit()

        # Emit event
        from app.plugins.events import emit

        emit(
            "iemis.import_completed",
            school_id=str(g.school_id),
            format=format_code,
            imported=result["imported"],
            errors=result["errors"],
        )

        return created_response(log.to_dict())

    except StudentCapExceededError as exc:
        db.session.rollback()
        log.status = "failed"
        log.errors = [{"error": str(exc)}]
        log.completed_at = datetime.now(UTC)
        db.session.commit()
        return error_response(str(exc), 403)

    except (ValueError, RuntimeError) as exc:
        log.status = "failed"
        log.errors = [{"error": str(exc)}]
        log.completed_at = datetime.now(UTC)
        db.session.commit()
        return error_response(str(exc), 422)

    except Exception as exc:
        log.status = "failed"
        log.errors = [{"error": f"Unexpected error: {exc}"}]
        log.completed_at = datetime.now(UTC)
        db.session.commit()
        return error_response("Import failed due to an internal error", 500)


@iemis_importer_bp.route("/history", methods=["GET"])
@jwt_required()
@school_required
@plugin_required("iemis_importer")
def import_history():
    """List all IEMIS imports for this school."""
    from app.utils.pagination import paginate

    q = IemisImportLog.query.filter_by(school_id=g.school_id, is_deleted=False)
    items, meta = paginate(q.order_by(IemisImportLog.created_at.desc()))
    return success_response([i.to_dict() for i in items], meta={"pagination": meta})


@iemis_importer_bp.route("/history/<uuid:log_id>", methods=["GET"])
@jwt_required()
@school_required
@plugin_required("iemis_importer")
def import_history_detail(log_id):
    """Get detail of a single import job."""
    log = IemisImportLog.query.filter_by(
        id=log_id, school_id=g.school_id, is_deleted=False
    ).first()
    if not log:
        return error_response("Import log not found", 404)
    return success_response(log.to_dict())
